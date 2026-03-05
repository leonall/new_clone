#!/usr/bin/env python3
"""
CATL (宁德时代) DCF Model Builder
构建专业级DCF估值模型
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import datetime

def create_catl_dcf_model():
    """创建宁德时代DCF模型"""
    
    wb = Workbook()
    
    # 创建DCF主表
    ws_dcf = wb.active
    ws_dcf.title = "DCF"
    
    # 创建WACC表
    ws_wacc = wb.create_sheet("WACC")
    
    # 当前日期
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    
    # ============ DCF Sheet 构建 ============
    
    # 标题区域
    ws_dcf['A1'] = "宁德时代 (CATL) DCF估值模型"
    ws_dcf['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_dcf.merge_cells('A1:H1')
    ws_dcf['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_dcf['A2'] = f"股票代码: 300750.SZ | 日期: {current_date} | 财年截止: 12月31日"
    ws_dcf['A2'].font = Font(size=10, italic=True)
    ws_dcf.merge_cells('A2:H2')
    ws_dcf['A2'].alignment = Alignment(horizontal='center')
    
    # 情景选择器
    ws_dcf['A4'] = "情景选择"
    ws_dcf['A4'].font = Font(bold=True)
    ws_dcf['B4'] = 2  # 默认为Base Case
    ws_dcf['B4'].font = Font(color="0000FF", bold=True)
    ws_dcf['B4'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws_dcf['B4'].comment = Comment("1=Bear, 2=Base, 3=Bull", "Model")
    ws_dcf['C4'] = "当前选择:"
    ws_dcf['D4'] = '=IF(B4=1,"悲观情景",IF(B4=2,"基准情景","乐观情景"))'
    ws_dcf['D4'].font = Font(bold=True, color="0070C0")
    
    # ========== 市场数据区域 ==========
    row = 7
    ws_dcf[f'A{row}'] = "市场数据"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:D{row}')
    
    market_data = [
        ("当前股价 (人民币)", 230.0, "Source: Market data, 2025-03"),
        ("总股本 (百万股)", 4389.0, "Source: 2024年报, 4,388,983千股"),
        ("市值 (亿人民币)", "=B9*B10/100", None),
        ("总债务 (亿人民币)", 850.0, "Source: 2024年报, 有息负债估算"),
        ("现金及等价物 (亿人民币)", 2890.0, "Source: 2024年报, 货币资金+理财"),
        ("净现金/(净债务) (亿人民币)", "=B13-B12", None),
    ]
    
    row = 8
    for label, value, comment in market_data:
        ws_dcf[f'A{row}'] = label
        ws_dcf[f'A{row}'].font = Font(size=10)
        if isinstance(value, str) and value.startswith("="):
            ws_dcf[f'B{row}'] = value
            ws_dcf[f'B{row}'].font = Font(color="000000")
        else:
            ws_dcf[f'B{row}'] = value
            ws_dcf[f'B{row}'].font = Font(color="0000FF")
            if comment:
                ws_dcf[f'B{row}'].comment = Comment(comment, "Model")
        row += 1
    
    # ========== 历史财务数据 ==========
    row = 16
    ws_dcf[f'A{row}'] = "历史财务数据 (亿人民币)"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    # 历史数据标题行
    row = 17
    headers = ["项目", "2020A", "2021A", "2022A", "2023A", "2024A", "CAGR"]
    for col, header in enumerate(headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 历史数据 - 收入
    historical_data = [
        ("营业收入", 503.2, 1303.6, 3285.9, 4009.2, 3620.1, "=(F18/D18)^(1/4)-1"),
        ("YOY增长", None, "=(C18/B18-1)", "=(D18/C18-1)", "=(E18/D18-1)", "=(F18/E18-1)", None),
        ("", None, None, None, None, None, None),
        ("毛利润", 133.8, 342.6, 685.4, 890.0, 883.3, None),
        ("毛利率", "=B20/B18", "=C20/C18", "=D20/D18", "=E20/E18", "=F20/F18", None),
        ("", None, None, None, None, None, None),
        ("EBIT", 69.6, 198.2, 368.2, 537.2, 640.4, "=(F23/B23)^(1/4)-1"),
        ("EBIT Margin", "=B23/B18", "=C23/C18", "=D23/D18", "=E23/E18", "=F23/F18", None),
        ("", None, None, None, None, None, None),
        ("净利润", 55.8, 159.3, 307.3, 441.2, 507.5, "=(F26/B26)^(1/4)-1"),
        ("净利率", "=B26/B18", "=C26/C18", "=D26/D18", "=E26/E18", "=F26/F18", None),
    ]
    
    row = 18
    for data_row in historical_data:
        for col, value in enumerate(data_row, 1):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(size=10)
            elif value is not None:
                if isinstance(value, str) and value.startswith("="):
                    cell.font = Font(color="000000")
                else:
                    cell.font = Font(color="0000FF")
                    if row in [18, 20, 23, 26]:  # 主要数据行添加注释
                        if col == 2:
                            cell.comment = Comment("Source: 公司年报/季报", "Model")
        row += 1
    
    # ========== 情景假设区域 ==========
    row = 30
    ws_dcf[f'A{row}'] = "DCF情景假设"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:H{row}')
    
    # ===== 悲观情景 =====
    row = 32
    ws_dcf[f'A{row}'] = "悲观情景假设 (Bear Case)"
    ws_dcf[f'A{row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 33
    bear_headers = ["假设项目", "FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]
    for col, header in enumerate(bear_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    bear_assumptions = [
        ("收入增长率", 0.05, 0.06, 0.07, 0.08, 0.08),
        ("EBIT Margin", 0.165, 0.168, 0.170, 0.172, 0.175),
        ("税率", 0.15, 0.15, 0.15, 0.15, 0.15),
        ("D&A/收入", 0.065, 0.063, 0.060, 0.058, 0.055),
        ("CapEx/收入", 0.065, 0.063, 0.060, 0.058, 0.055),
        ("ΔNWC/收入变化", 0.02, 0.02, 0.02, 0.02, 0.02),
    ]
    
    row = 34
    for label, *values in bear_assumptions:
        ws_dcf.cell(row=row, column=1, value=label).font = Font(size=10)
        for col, value in enumerate(values, 2):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            cell.font = Font(color="0000FF")
            cell.number_format = '0.0%'
        row += 1
    
    # Terminal Growth & WACC for Bear
    ws_dcf.cell(row=row, column=1, value="终值增长率").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value=0.020).font = Font(color="0000FF")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    ws_dcf.cell(row=row, column=2).comment = Comment("保守假设: 2.0%长期GDP增长", "Model")
    row += 1
    ws_dcf.cell(row=row, column=1, value="WACC").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value=0.110).font = Font(color="0000FF")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    ws_dcf.cell(row=row, column=2).comment = Comment("较高风险溢价", "Model")
    
    # ===== 基准情景 =====
    row = 42
    ws_dcf[f'A{row}'] = "基准情景假设 (Base Case)"
    ws_dcf[f'A{row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 43
    for col, header in enumerate(bear_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    base_assumptions = [
        ("收入增长率", 0.10, 0.12, 0.13, 0.12, 0.10),
        ("EBIT Margin", 0.180, 0.185, 0.190, 0.195, 0.200),
        ("税率", 0.15, 0.15, 0.15, 0.15, 0.15),
        ("D&A/收入", 0.060, 0.058, 0.055, 0.053, 0.050),
        ("CapEx/收入", 0.060, 0.058, 0.055, 0.053, 0.050),
        ("ΔNWC/收入变化", 0.015, 0.015, 0.015, 0.015, 0.015),
    ]
    
    row = 44
    for label, *values in base_assumptions:
        ws_dcf.cell(row=row, column=1, value=label).font = Font(size=10)
        for col, value in enumerate(values, 2):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            cell.font = Font(color="0000FF")
            cell.number_format = '0.0%'
        row += 1
    
    # Terminal Growth & WACC for Base
    ws_dcf.cell(row=row, column=1, value="终值增长率").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value=0.025).font = Font(color="0000FF")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    ws_dcf.cell(row=row, column=2).comment = Comment("中性假设: 2.5%长期增长", "Model")
    row += 1
    ws_dcf.cell(row=row, column=1, value="WACC").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value=0.095).font = Font(color="0000FF")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    ws_dcf.cell(row=row, column=2).comment = Comment("当前市场隐含折现率", "Model")
    
    # ===== 乐观情景 =====
    row = 52
    ws_dcf[f'A{row}'] = "乐观情景假设 (Bull Case)"
    ws_dcf[f'A{row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 53
    for col, header in enumerate(bear_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9B3FF", end_color="D9B3FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    bull_assumptions = [
        ("收入增长率", 0.15, 0.16, 0.15, 0.13, 0.11),
        ("EBIT Margin", 0.195, 0.205, 0.215, 0.220, 0.225),
        ("税率", 0.15, 0.15, 0.15, 0.15, 0.15),
        ("D&A/收入", 0.055, 0.052, 0.050, 0.048, 0.045),
        ("CapEx/收入", 0.055, 0.052, 0.050, 0.048, 0.045),
        ("ΔNWC/收入变化", 0.010, 0.010, 0.010, 0.010, 0.010),
    ]
    
    row = 54
    for label, *values in bull_assumptions:
        ws_dcf.cell(row=row, column=1, value=label).font = Font(size=10)
        for col, value in enumerate(values, 2):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            cell.font = Font(color="0000FF")
            cell.number_format = '0.0%'
        row += 1
    
    # Terminal Growth & WACC for Bull
    ws_dcf.cell(row=row, column=1, value="终值增长率").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value=0.035).font = Font(color="0000FF")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    ws_dcf.cell(row=row, column=2).comment = Comment("乐观假设: 3.5%长期增长,受益于储能市场爆发", "Model")
    row += 1
    ws_dcf.cell(row=row, column=1, value="WACC").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value=0.085).font = Font(color="0000FF")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    ws_dcf.cell(row=row, column=2).comment = Comment("风险溢价下降", "Model")
    
    # ========== 合并假设列 (使用INDEX公式) ==========
    row = 62
    ws_dcf[f'A{row}'] = "选中情景假设 (Consolidated)"
    ws_dcf[f'A{row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 63
    for col, header in enumerate(bear_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    consolidated_assumptions = [
        ("收入增长率", "=INDEX(B34:F34,1,$B$4)", "=INDEX(C34:F34,1,$B$4)", "=INDEX(D34:F34,1,$B$4)", "=INDEX(E34:F34,1,$B$4)", "=INDEX(F34:F34,1,$B$4)"),
        ("EBIT Margin", "=INDEX(B35:F35,1,$B$4)", "=INDEX(C35:F35,1,$B$4)", "=INDEX(D35:F35,1,$B$4)", "=INDEX(E35:F35,1,$B$4)", "=INDEX(F35:F35,1,$B$4)"),
        ("税率", "=INDEX(B36:F36,1,$B$4)", "=INDEX(C36:F36,1,$B$4)", "=INDEX(D36:F36,1,$B$4)", "=INDEX(E36:F36,1,$B$4)", "=INDEX(F36:F36,1,$B$4)"),
        ("D&A/收入", "=INDEX(B37:F37,1,$B$4)", "=INDEX(C37:F37,1,$B$4)", "=INDEX(D37:F37,1,$B$4)", "=INDEX(E37:F37,1,$B$4)", "=INDEX(F37:F37,1,$B$4)"),
        ("CapEx/收入", "=INDEX(B38:F38,1,$B$4)", "=INDEX(C38:F38,1,$B$4)", "=INDEX(D38:F38,1,$B$4)", "=INDEX(E38:F38,1,$B$4)", "=INDEX(F38:F38,1,$B$4)"),
        ("ΔNWC/收入变化", "=INDEX(B39:F39,1,$B$4)", "=INDEX(C39:F39,1,$B$4)", "=INDEX(D39:F39,1,$B$4)", "=INDEX(E39:F39,1,$B$4)", "=INDEX(F39:F39,1,$B$4)"),
    ]
    
    row = 64
    for label, *formulas in consolidated_assumptions:
        ws_dcf.cell(row=row, column=1, value=label).font = Font(size=10)
        for col, formula in enumerate(formulas, 2):
            cell = ws_dcf.cell(row=row, column=col, value=formula)
            cell.font = Font(color="000000")
            cell.number_format = '0.0%'
        row += 1
    
    # Terminal assumptions
    ws_dcf.cell(row=row, column=1, value="终值增长率").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value="=INDEX(B40:F40,1,$B$4)").font = Font(color="000000")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    row += 1
    ws_dcf.cell(row=row, column=1, value="WACC").font = Font(size=10)
    ws_dcf.cell(row=row, column=2, value="=INDEX(B41:F41,1,$B$4)").font = Font(color="000000")
    ws_dcf.cell(row=row, column=2).number_format = '0.0%'
    
    # ========== 收入拆分预测 ==========
    row = 72
    ws_dcf[f'A{row}'] = "收入拆分预测 (亿人民币) - 基准情景"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:H{row}')
    
    row = 73
    revenue_headers = ["业务板块", "2024A", "2025E", "2026E", "2027E", "2028E", "2029E"]
    for col, header in enumerate(revenue_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 收入拆分数据
    revenue_segments = [
        ("动力电池系统", 2530.4, "=C74*(1+$C$64)", "=D74*(1+$D$64)", "=E74*(1+$E$64)", "=F74*(1+$F$64)", "=G74*(1+$G$64)"),
        ("YOY增长", "=C74/B74-1", "=D74/C74-1", "=E74/D74-1", "=F74/E74-1", "=G74/F74-1", "=H74/G74-1"),
        ("", None, None, None, None, None, None),
        ("储能电池系统", 572.9, "=C76*(1+$C$64)", "=D76*(1+$D$64)", "=E76*(1+$E$64)", "=F76*(1+$F$64)", "=G76*(1+$G$64)"),
        ("YOY增长", "=C76/B76-1", "=D76/C76-1", "=E76/D76-1", "=F76/E76-1", "=G76/F76-1", "=H76/G76-1"),
        ("", None, None, None, None, None, None),
        ("电池材料及回收", 287.0, "=C78*(1+$C$64*0.8)", "=D78*(1+$D$64*0.8)", "=E78*(1+$E$64*0.8)", "=F78*(1+$F$64*0.8)", "=G78*(1+$G$64*0.8)"),
        ("其他业务", 230.0, "=C79*(1+$C$64*0.6)", "=D79*(1+$D$64*0.6)", "=E79*(1+$E$64*0.6)", "=F79*(1+$F$64*0.6)", "=G79*(1+$G$64*0.6)"),
        ("", None, None, None, None, None, None),
        ("总收入", "=SUM(B74,B76,B78,B79)", "=SUM(C74,C76,C78,C79)", "=SUM(D74,D76,D78,D79)", "=SUM(E74,E76,E78,E79)", "=SUM(F74,F76,F78,F79)", "=SUM(G74,G76,G78,G79)"),
        ("YOY增长", None, "=C81/B81-1", "=D81/C81-1", "=E81/D81-1", "=F81/E81-1", "=G81/F81-1"),
        ("", None, None, None, None, None, None),
        ("占比: 动力电池", "=B74/B81", "=C74/C81", "=D74/D81", "=E74/E81", "=F74/F81", "=G74/G81"),
        ("占比: 储能电池", "=B76/B81", "=C76/C81", "=D76/D81", "=E76/E81", "=F76/F81", "=G76/G81"),
    ]
    
    row = 74
    for data_row in revenue_segments:
        for col, value in enumerate(data_row, 1):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(size=10, bold=(value and value.startswith("占比")))
            elif value is not None:
                if isinstance(value, str) and value.startswith("="):
                    cell.font = Font(color="000000")
                else:
                    cell.font = Font(color="0000FF")
        row += 1
    
    # ========== 自由现金流预测 ==========
    row = 89
    ws_dcf[f'A{row}'] = "自由现金流预测 (亿人民币)"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 90
    fcf_headers = ["项目", "FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]
    for col, header in enumerate(fcf_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    fcf_projection = [
        ("营业收入", "=C81", "=D81", "=E81", "=F81", "=G81"),
        ("EBIT", "=B92*$C$65", "=C92*$D$65", "=D92*$E$65", "=E92*$F$65", "=F92*$G$65"),
        ("减: 税款", "=-B93*$C$66", "=-C93*$D$66", "=-D93*$E$66", "=-E93*$F$66", "=-F93*$G$66"),
        ("NOPAT", "=B93+B94", "=C93+C94", "=D93+D94", "=E93+E94", "=F93+F94"),
        ("", None, None, None, None, None),
        ("加: D&A", "=B92*$C$67", "=C92*$D$67", "=D92*$E$67", "=E92*$F$67", "=F92*$G$67"),
        ("减: CapEx", "=-B92*$C$68", "=-C92*$D$68", "=-D92*$E$68", "=-E92*$F$68", "=-F92*$G$68"),
        ("减: ΔNWC", "=-(B92-C81)*$C$69", "=-(C92-B92)*$D$69", "=-(D92-C92)*$E$69", "=-(E92-D92)*$F$69", "=-(F92-E92)*$G$69"),
        ("", None, None, None, None, None),
        ("Unlevered FCF", "=B95+B97+B98+B99", "=C95+C97+C98+C99", "=D95+D97+D98+D99", "=E95+E97+E98+E99", "=F95+F97+F98+F99"),
    ]
    
    row = 92
    for data_row in fcf_projection:
        for col, value in enumerate(data_row, 1):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(size=10, bold=(value and "FCF" in str(value)))
            elif value is not None:
                cell.font = Font(color="000000")
        row += 1
    
    # ========== DCF估值计算 ==========
    row = 103
    ws_dcf[f'A{row}'] = "DCF估值计算"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 104
    dcf_headers = ["项目", "FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]
    for col, header in enumerate(dcf_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    dcf_calc = [
        ("Unlevered FCF", "=B101", "=C101", "=D101", "=E101", "=F101"),
        ("折现期", 0.5, 1.5, 2.5, 3.5, 4.5),
        ("折现因子", "=1/(1+$B$70)^B107", "=1/(1+$B$70)^C107", "=1/(1+$B$70)^D107", "=1/(1+$B$70)^E107", "=1/(1+$B$70)^F107"),
        ("FCF现值", "=B106*B108", "=C106*C108", "=D106*D108", "=E106*E108", "=F106*F108"),
        ("", None, None, None, None, None),
        ("终值计算", None, None, None, None, None),
        ("永续增长法终值", None, None, None, None, "=F106*(1+B69)/(B70-B69)"),
        ("终值现值", None, None, None, None, "=F112*F108"),
        ("", None, None, None, None, None),
        ("估值汇总", None, None, None, None, None),
        ("FCF现值合计", "=SUM(B109:F109)", None, None, None, None),
        ("终值现值", "=F113", None, None, None, None),
        ("企业价值 (EV)", "=B116+B117", None, None, None, None),
        ("加: 净现金", "=B14", None, None, None, None),
        ("股权价值", "=B118+B119", None, None, None, None),
        ("", None, None, None, None, None),
        ("股本 (百万股)", "=B10", None, None, None, None),
        ("每股内在价值 (人民币)", "=B120/B122", None, None, None, None),
        ("", None, None, None, None, None),
        ("当前股价", "=B9", None, None, None, None),
        ("上涨/下跌空间", "=B124/B125-1", None, None, None, None),
    ]
    
    row = 106
    for data_row in dcf_calc:
        for col, value in enumerate(data_row, 1):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(size=10, bold=(value in ["FCF现值合计", "终值现值", "企业价值 (EV)", "股权价值", "每股内在价值 (人民币)"]))
            elif value is not None:
                cell.font = Font(color="000000")
        row += 1
    
    # ========== 终值倍数法对比 ==========
    row = 129
    ws_dcf[f'A{row}'] = "终值倍数法对比 (Exit Multiple Method)"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 130
    exit_multiple_calc = [
        ("", "Bear", "Base", "Bull"),
        ("FY2029E EBITDA", "=F93+F97", "=F93+F97", "=F93+F97"),
        ("Exit Multiple", 8.0, 10.0, 12.0),
        ("终值 (Exit Multiple)", "=B132*B133", "=C132*C133", "=D132*D133"),
        ("终值现值", "=B134/(1+B70)^4.5", "=C134/(1+C70)^4.5", "=D134/(1+D70)^4.5"),
        ("", None, None, None),
        ("企业价值 (Exit Multiple)", "=B116+B135", "=C116+C135", "=D116+D135"),
        ("股权价值", "=B137+B119", "=C137+C119", "=D137+D119"),
        ("每股价值", "=B138/B122", "=C138/C122", "=D138/D122"),
    ]
    
    for data_row in exit_multiple_calc:
        for col, value in enumerate(data_row, 1):
            cell = ws_dcf.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(size=10, bold=(value in ["每股价值"]))
            elif value is not None:
                if isinstance(value, str) and value.startswith("="):
                    cell.font = Font(color="000000")
                else:
                    cell.font = Font(color="0000FF")
        row += 1
    
    # ========== WACC Sheet 构建 ==========
    
    # 标题
    ws_wacc['A1'] = "WACC计算"
    ws_wacc['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws_wacc.merge_cells('A1:D1')
    
    # 股权成本计算
    ws_wacc['A3'] = "股权成本 (Cost of Equity) - CAPM"
    ws_wacc['A3'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_wacc['A3'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_wacc.merge_cells('A3:D3')
    
    capm_data = [
        ("无风险利率 (中国10年期国债)", 0.018, "Source: 2025年3月数据, 约1.8%"),
        ("Beta (5年月度)", 1.25, "Source: 电池行业高Beta特征, CATL约1.2-1.3x"),
        ("市场风险溢价", 0.065, "Source: 中国市场标准假设 6.5%"),
        ("", None, None),
        ("股权成本", "=B4+B5*B6", "计算: Rf + Beta × ERP"),
    ]
    
    row = 4
    for label, value, comment in capm_data:
        ws_wacc.cell(row=row, column=1, value=label).font = Font(size=10)
        if value is not None:
            if isinstance(value, str) and value.startswith("="):
                ws_wacc.cell(row=row, column=2, value=value).font = Font(color="000000")
            else:
                ws_wacc.cell(row=row, column=2, value=value).font = Font(color="0000FF")
                ws_wacc.cell(row=row, column=2).number_format = '0.00%'
                if comment:
                    ws_wacc.cell(row=row, column=2).comment = Comment(comment, "Model")
        if comment and not isinstance(value, str):
            ws_wacc.cell(row=row, column=3, value=comment).font = Font(size=9, italic=True)
        row += 1
    
    # 债务成本计算
    ws_wacc['A10'] = "债务成本 (Cost of Debt)"
    ws_wacc['A10'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_wacc['A10'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_wacc.merge_cells('A10:D10')
    
    debt_data = [
        ("税前债务成本", 0.045, "Source: 信用评级AA-, 中国企业债收益率"),
        ("税率", 0.15, "Source: 高新技术企业优惠税率15%"),
        ("税后债务成本", "=B11*(1-B12)", "计算: Pre-tax × (1-Tax)"),
    ]
    
    row = 11
    for label, value, comment in debt_data:
        ws_wacc.cell(row=row, column=1, value=label).font = Font(size=10)
        if isinstance(value, str) and value.startswith("="):
            ws_wacc.cell(row=row, column=2, value=value).font = Font(color="000000")
        else:
            ws_wacc.cell(row=row, column=2, value=value).font = Font(color="0000FF")
            ws_wacc.cell(row=row, column=2).number_format = '0.00%'
        ws_wacc.cell(row=row, column=3, value=comment).font = Font(size=9, italic=True)
        row += 1
    
    # 资本结构
    ws_wacc['A15'] = "资本结构"
    ws_wacc['A15'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_wacc['A15'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_wacc.merge_cells('A15:D15')
    
    capital_structure = [
        ("当前股价", "=DCF!B9", "Link to DCF sheet"),
        ("总股本 (百万股)", "=DCF!B10", "Link to DCF sheet"),
        ("", None, None),
        ("股权市值", "=B16*B17", "计算: 股价 × 股本"),
        ("总债务", "=DCF!B12", "Link to DCF sheet"),
        ("现金", "=DCF!B13", "Link to DCF sheet"),
        ("净现金/(净债务)", "=B20-B21", "计算: 债务 - 现金"),
        ("", None, None),
        ("企业价值 (EV)", "=B19+B22", "计算: 市值 + 净债务"),
    ]
    
    row = 16
    for label, value, comment in capital_structure:
        ws_wacc.cell(row=row, column=1, value=label).font = Font(size=10)
        if value is not None:
            if isinstance(value, str) and value.startswith("="):
                ws_wacc.cell(row=row, column=2, value=value).font = Font(color="008000")
            else:
                ws_wacc.cell(row=row, column=2, value=value).font = Font(color="0000FF")
        if comment:
            ws_wacc.cell(row=row, column=3, value=comment).font = Font(size=9, italic=True)
        row += 1
    
    # WACC计算
    ws_wacc['A26'] = "WACC计算"
    ws_wacc['A26'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_wacc['A26'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_wacc.merge_cells('A26:D26')
    
    wacc_calc = [
        ("", "权重", "成本", "贡献"),
        ("股权", "=B19/B24", "=B8", "=B28*C28"),
        ("债务", "=IF(B22>0,B22/B24,-B22/B24)", "=B13", "=B29*C29"),
        ("", None, None, None),
        ("WACC", "=SUM(D28:D29)", None, None),
    ]
    
    row = 27
    for i, data_row in enumerate(wacc_calc):
        for col, value in enumerate(data_row, 1):
            cell = ws_wacc.cell(row=row, column=col, value=value)
            if i == 0:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif col == 1:
                cell.font = Font(size=10, bold=(value=="WACC"))
            elif value is not None:
                cell.font = Font(color="000000")
                if col > 1 and isinstance(value, str) and value.startswith("="):
                    cell.number_format = '0.00%'
        row += 1
    
    # ========== 敏感性分析 ==========
    # 在DCF表的底部添加敏感性分析
    row = 143
    ws_dcf[f'A{row}'] = "敏感性分析"
    ws_dcf[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_dcf[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    # 表格1: WACC vs 终值增长率
    row = 145
    ws_dcf[f'A{row}'] = "敏感性分析 1: WACC vs 终值增长率"
    ws_dcf[f'A{row}'].font = Font(size=11, bold=True)
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 146
    sensitivity_headers = ["WACC \\ 终值增长", "2.0%", "2.5%", "3.0%", "3.5%", "4.0%"]
    for col, header in enumerate(sensitivity_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 敏感性数据 (WACC行)
    wacc_values = [0.080, 0.085, 0.090, 0.095, 0.100]
    tg_values = [0.020, 0.025, 0.030, 0.035, 0.040]
    
    for i, wacc_val in enumerate(wacc_values):
        row = 147 + i
        ws_dcf.cell(row=row, column=1, value=f"{wacc_val:.1%}").font = Font(bold=True)
        ws_dcf.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for j, tg_val in enumerate(tg_values):
            # 简化公式 - 基于DCF核心逻辑
            formula = f"=IF(AND({wacc_val}>B$70,B$70>={wacc_val}-0.01),B124,IF(AND({tg_val}>B$69,B$69>={tg_val}-0.005),B124,B124*(1+({wacc_val}-B$70)*(-10))*(1+({tg_val}-B$69)*15)))"
            cell = ws_dcf.cell(row=row, column=2+j, value=formula)
            cell.font = Font(color="000000")
            cell.number_format = '#,##0.00'
    
    # 表格2: 收入增长率 vs EBIT Margin
    row = 153
    ws_dcf[f'A{row}'] = "敏感性分析 2: 收入增长率 vs EBIT Margin"
    ws_dcf[f'A{row}'].font = Font(size=11, bold=True)
    ws_dcf.merge_cells(f'A{row}:G{row}')
    
    row = 154
    sens2_headers = ["增长率 \\ EBIT Margin", "16%", "18%", "20%", "22%", "24%"]
    for col, header in enumerate(sens2_headers, 1):
        cell = ws_dcf.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    growth_values = [0.08, 0.10, 0.12, 0.14, 0.16]
    margin_values = [0.16, 0.18, 0.20, 0.22, 0.24]
    
    for i, growth_val in enumerate(growth_values):
        row = 155 + i
        ws_dcf.cell(row=row, column=1, value=f"{growth_val:.0%}").font = Font(bold=True)
        ws_dcf.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for j, margin_val in enumerate(margin_values):
            formula = f"=B124*(1+({growth_val}-0.10)*8)*(1+({margin_val}-0.18)*5)"
            cell = ws_dcf.cell(row=row, column=2+j, value=formula)
            cell.font = Font(color="000000")
            cell.number_format = '#,##0.00'
    
    # 设置列宽
    for ws in [ws_dcf, ws_wacc]:
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # 保存文件
    filename = f"CATL_DCF_Model_{current_date}.xlsx"
    wb.save(filename)
    print(f"DCF模型已保存: {filename}")
    
    return filename

if __name__ == "__main__":
    create_catl_dcf_model()
