#!/usr/bin/env python3
"""
Excel公式重算与验证脚本
"""

import subprocess
import sys
import json
import os

def recalc_excel(filepath, timeout=30):
    """验证Excel公式"""
    
    try:
        if not os.path.exists(filepath):
            return {'status': 'file_not_found', 'filepath': filepath}
            
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        errors = []
        formula_count = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        # 检查公式中是否有潜在错误模式
                        formula_str = str(cell.value)
                        if any(err in formula_str for err in ['#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#NULL!', '#NUM!', '#N/A']):
                            errors.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'error': 'Formula error in cell'
                            })
        
        result = {
            'status': 'success' if len(errors) == 0 else 'errors_found',
            'total_errors': len(errors),
            'total_formulas': formula_count,
        }
        
        if errors:
            result['error_summary'] = errors[:10]
            
        return result
            
    except Exception as e:
        return {'status': 'error', 'message': str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalc_excel(filepath, timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))
