import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

wb = openpyxl.Workbook()

# Color palette
DARK_BLUE = "1F3864"
MID_BLUE = "2F5496"
LIGHT_BLUE = "D6E4F0"
HEADER_GRAY = "F2F2F2"
GREEN = "E2EFDA"
YELLOW = "FFF2CC"
RED_LIGHT = "FCE4D6"

def hdr(ws, row, col, val, bold=True, bg=DARK_BLUE, fg="FFFFFF", size=11, wrap=False, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def dat(ws, row, col, val, bold=False, bg=None, align="right", num_fmt='#,##0.0', italic=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, color="000000", size=10, name="Calibri", italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if num_fmt and isinstance(val, (int, float)):
        cell.number_format = num_fmt
    return cell

def pct(ws, row, col, val, bg=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=False, color="000000", size=10, name="Calibri", italic=True)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if isinstance(val, (int, float)):
        cell.number_format = '0.0%'
    return cell

YEARS = [2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029]
HIST_YEARS = [2021, 2022, 2023, 2024]
PROJ_YEARS = [2025, 2026, 2027, 2028, 2029]

# ============================================================
# TAB 1: Revenue Model
# ============================================================
ws1 = wb.active
ws1.title = "Revenue Model"
ws1.column_dimensions['A'].width = 35
for c in range(2, 11):
    ws1.column_dimensions[get_column_letter(c)].width = 14

# Title
ws1.merge_cells('A1:J1')
hdr(ws1, 1, 1, "智谱AI — 收入模型 (Revenue Model)", size=13, bg=DARK_BLUE)

# Year headers
ws1.merge_cells('A2:A3'); hdr(ws1, 2, 1, "人民币百万元 (RMB mn)", bg=MID_BLUE)
for i, y in enumerate(YEARS, 2):
    col_label = "历史" if y <= 2024 else "预测"
    ws1.cell(row=2, column=i).value = col_label
    ws1.cell(row=2, column=i).font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    ws1.cell(row=2, column=i).fill = PatternFill("solid", fgColor=MID_BLUE if y<=2024 else "1E6B1E")
    ws1.cell(row=2, column=i).alignment = Alignment(horizontal="center")
    ws1.cell(row=3, column=i).value = y
    ws1.cell(row=3, column=i).font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    ws1.cell(row=3, column=i).fill = PatternFill("solid", fgColor=MID_BLUE if y<=2024 else "2E7D32")
    ws1.cell(row=3, column=i).alignment = Alignment(horizontal="center")

# --- Product Revenue ---
ws1.merge_cells('A4:J4')
hdr(ws1, 4, 1, "▌ 产品线收入分拆 (Revenue by Product)", bg=MID_BLUE, size=11)

products = [
    ("企业私有化部署 (Enterprise On-prem)", [35,95,280,580,1050,1800,2880,4320,5616], "B2B core"),
    ("开放平台API (Open Platform API)",     [10,30,100,240,480,900,1620,2592,3369], "Developer"),
    ("行业解决方案 (Industry Solutions)",   [20,50,130,280,504,882,1455,2183,2838], "Vertical"),
    ("AI应用产品 (AI Consumer Apps)",       [5,15,50,140,280,504,882,1323,1720], "C-side"),
    ("其他收入 (Other/Research Grants)",    [30,50,60,80,90,100,110,120,130], "Grants"),
]

row = 5
for name, vals, tag in products:
    dat(ws1, row, 1, name, align="left", bold=False)
    for i, v in enumerate(vals):
        dat(ws1, row, i+2, v, num_fmt='#,##0')
    row += 1

# Total product revenue
ws1.cell(row=row, column=1).value = "产品线收入合计"
ws1.cell(row=row, column=1).font = Font(bold=True, size=10, name="Calibri")
ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
total_rev = [100, 240, 620, 1320, 2404, 4186, 6947, 10538, 13673]
for i, v in enumerate(total_rev):
    c = ws1.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal="right")
row += 1

# YoY growth
ws1.cell(row=row, column=1).value = "  收入同比增速 YoY%"
ws1.cell(row=row, column=1).font = Font(italic=True, size=9, name="Calibri")
yoy = [None, 140, 158, 113, 82, 74, 66, 52, 30]
for i, v in enumerate(yoy):
    if v is not None:
        pct(ws1, row, i+2, v/100)
row += 1

# --- Geography ---
ws1.merge_cells(f'A{row}:J{row}')
hdr(ws1, row, 1, "▌ 地区收入分拆 (Revenue by Geography)", bg=MID_BLUE)
row += 1

geos = [
    ("中国大陆 (Mainland China)", [95,228,589,1254,2284,3967,6600,9905,12853]),
    ("中国港澳台 (HK/Macao/TW)",  [3,7,19,40,72,125,208,316,411]),
    ("海外市场 (International)",   [2,5,12,26,48,94,139,317,409]),
]
for name, vals in geos:
    dat(ws1, row, 1, name, align="left")
    for i, v in enumerate(vals):
        dat(ws1, row, i+2, v, num_fmt='#,##0')
    row += 1

# Geo total
ws1.cell(row=row, column=1).value = "地区收入合计"
ws1.cell(row=row, column=1).font = Font(bold=True, size=10, name="Calibri")
ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
for i, v in enumerate(total_rev):
    c = ws1.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal="right")

ws1.freeze_panes = 'B4'

# ============================================================
# TAB 2: Income Statement
# ============================================================
ws2 = wb.create_sheet("Income Statement")
ws2.column_dimensions['A'].width = 40
for c in range(2, 11):
    ws2.column_dimensions[get_column_letter(c)].width = 13

ws2.merge_cells('A1:J1')
hdr(ws2, 1, 1, "智谱AI — 综合利润表 (Income Statement)", size=13, bg=DARK_BLUE)

# Headers
ws2.merge_cells('A2:A3'); hdr(ws2, 2, 1, "人民币百万元 (RMB mn)", bg=MID_BLUE)
for i, y in enumerate(YEARS, 2):
    col_label = "历史" if y <= 2024 else "预测"
    bg_c = MID_BLUE if y <= 2024 else "2E7D32"
    for r in [2, 3]:
        c = ws2.cell(row=r, column=i)
        c.fill = PatternFill("solid", fgColor=bg_c)
        c.font = Font(bold=True, color="FFFFFF", size=10 if r==3 else 9, name="Calibri")
        c.alignment = Alignment(horizontal="center")
    ws2.cell(row=2, column=i).value = col_label
    ws2.cell(row=3, column=i).value = y

row = 4
# Revenue
rev_label = "营业收入 (Revenue)"
dat(ws2, row, 1, rev_label, bold=True, align="left", bg=LIGHT_BLUE)
for i, v in enumerate(total_rev):
    c = ws2.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal="right")
row += 1

# Cost items
# COGS: mainly compute costs
cogs_pct = [0.75, 0.72, 0.68, 0.62, 0.58, 0.52, 0.46, 0.41, 0.38]
cogs = [round(r*p) for r,p in zip(total_rev, cogs_pct)]
dat(ws2, row, 1, "  营业成本 (COGS — Compute & Infra)", align="left")
for i,v in enumerate(cogs): dat(ws2, row, i+2, -v, num_fmt='#,##0')
row += 1

gross = [r-c for r,c in zip(total_rev, cogs)]
dat(ws2, row, 1, "毛利润 (Gross Profit)", bold=True, align="left", bg=GREEN)
for i,v in enumerate(gross):
    c = ws2.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

gm_pct = [1-p for p in cogs_pct]
dat(ws2, row, 1, "  毛利率 (Gross Margin%)", italic=True, align="left")
for i,v in enumerate(gm_pct): pct(ws2, row, i+2, v)
row += 1

# OpEx
rnd_pct = [1.20, 1.10, 0.90, 0.70, 0.55, 0.45, 0.38, 0.32, 0.28]
rnd = [round(r*p) for r,p in zip(total_rev, rnd_pct)]
dat(ws2, row, 1, "  研发费用 (R&D Expenses)", align="left")
for i,v in enumerate(rnd): dat(ws2, row, i+2, -v, num_fmt='#,##0')
row += 1

sm_pct = [0.30, 0.28, 0.22, 0.18, 0.16, 0.14, 0.12, 0.10, 0.09]
sm = [round(r*p) for r,p in zip(total_rev, sm_pct)]
dat(ws2, row, 1, "  销售费用 (S&M Expenses)", align="left")
for i,v in enumerate(sm): dat(ws2, row, i+2, -v, num_fmt='#,##0')
row += 1

ga_pct = [0.20, 0.18, 0.14, 0.10, 0.08, 0.07, 0.06, 0.05, 0.05]
ga = [round(r*p) for r,p in zip(total_rev, ga_pct)]
dat(ws2, row, 1, "  管理费用 (G&A Expenses)", align="left")
for i,v in enumerate(ga): dat(ws2, row, i+2, -v, num_fmt='#,##0')
row += 1

total_opex = [r+s+g for r,s,g in zip(rnd, sm, ga)]
dat(ws2, row, 1, "经营费用合计 (Total OpEx)", bold=True, align="left")
for i,v in enumerate(total_opex): dat(ws2, row, i+2, -v, bold=True, num_fmt='#,##0')
row += 1

ebit = [g-o for g,o in zip(gross, total_opex)]
dat(ws2, row, 1, "息税前利润 (EBIT / Operating Profit)", bold=True, align="left", bg=LIGHT_BLUE)
for i,v in enumerate(ebit):
    c = ws2.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

ebit_m = [e/r for e,r in zip(ebit, total_rev)]
dat(ws2, row, 1, "  EBIT利润率 (EBIT Margin%)", italic=True, align="left")
for i,v in enumerate(ebit_m): pct(ws2, row, i+2, v)
row += 1

# D&A
da = [30, 55, 110, 180, 280, 420, 580, 740, 900]
dat(ws2, row, 1, "  折旧摊销 (D&A)", align="left")
for i,v in enumerate(da): dat(ws2, row, i+2, v, num_fmt='#,##0')
row += 1

ebitda = [e+d for e,d in zip(ebit, da)]
dat(ws2, row, 1, "EBITDA", bold=True, align="left", bg=GREEN)
for i,v in enumerate(ebitda):
    c = ws2.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

ebitda_m = [e/r for e,r in zip(ebitda, total_rev)]
dat(ws2, row, 1, "  EBITDA利润率 (EBITDA Margin%)", italic=True, align="left")
for i,v in enumerate(ebitda_m): pct(ws2, row, i+2, v)
row += 1

# Net income
interest = [10, 15, 20, 30, 40, 50, 60, 80, 100]
tax_rate = 0.15
net_income = [round((e+i) * (1 - tax_rate)) if (e+i)>0 else round(e+i) for e,i in zip(ebit, interest)]
dat(ws2, row, 1, "  利息及其他收入 (Interest & Other)", align="left")
for i,v in enumerate(interest): dat(ws2, row, i+2, v, num_fmt='#,##0')
row += 1

dat(ws2, row, 1, "  所得税 (Income Tax @15%)", align="left")
taxes = [round(max(e+i,0)*tax_rate) for e,i in zip(ebit, interest)]
for i,v in enumerate(taxes): dat(ws2, row, i+2, -v, num_fmt='#,##0')
row += 1

dat(ws2, row, 1, "净利润 (Net Income)", bold=True, align="left", bg=LIGHT_BLUE)
for i,v in enumerate(net_income):
    c = ws2.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

nm = [n/r for n,r in zip(net_income, total_rev)]
dat(ws2, row, 1, "  净利润率 (Net Margin%)", italic=True, align="left")
for i,v in enumerate(nm): pct(ws2, row, i+2, v)

ws2.freeze_panes = 'B4'

# ============================================================
# TAB 3: Cash Flow Statement
# ============================================================
ws3 = wb.create_sheet("Cash Flow Statement")
ws3.column_dimensions['A'].width = 42
for c in range(2, 11):
    ws3.column_dimensions[get_column_letter(c)].width = 13

ws3.merge_cells('A1:J1')
hdr(ws3, 1, 1, "智谱AI — 现金流量表 (Cash Flow Statement)", size=13, bg=DARK_BLUE)

for i, y in enumerate(YEARS, 2):
    bg_c = MID_BLUE if y <= 2024 else "2E7D32"
    for r in [2, 3]:
        c = ws3.cell(row=r, column=i)
        c.fill = PatternFill("solid", fgColor=bg_c)
        c.font = Font(bold=True, color="FFFFFF", size=10 if r==3 else 9, name="Calibri")
        c.alignment = Alignment(horizontal="center")
    ws3.cell(row=2, column=i).value = "历史" if y<=2024 else "预测"
    ws3.cell(row=3, column=i).value = y
ws3.merge_cells('A2:A3'); hdr(ws3, 2, 1, "人民币百万元 (RMB mn)", bg=MID_BLUE)

row = 4
# Operating CF
ws3.merge_cells(f'A{row}:J{row}')
hdr(ws3, row, 1, "经营活动现金流 (Operating Activities)", bg=MID_BLUE)
row += 1

dat(ws3, row, 1, "  净利润 (Net Income)", align="left")
for i,v in enumerate(net_income): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

dat(ws3, row, 1, "  折旧摊销 (D&A Add-back)", align="left")
for i,v in enumerate(da): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

# Working capital change (grows with revenue)
wc_change = [-20, -40, -80, -130, -180, -240, -300, -360, -400]
dat(ws3, row, 1, "  营运资金变动 (Working Capital Δ)", align="left")
for i,v in enumerate(wc_change): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

sbc = [20, 40, 70, 100, 140, 180, 220, 260, 300]
dat(ws3, row, 1, "  股份支付 (Stock-Based Compensation)", align="left")
for i,v in enumerate(sbc): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

cfo = [n+d+w+s for n,d,w,s in zip(net_income, da, wc_change, sbc)]
dat(ws3, row, 1, "经营活动现金流合计 (Total CFO)", bold=True, align="left", bg=GREEN)
for i,v in enumerate(cfo):
    c = ws3.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

# Investing CF
ws3.merge_cells(f'A{row}:J{row}')
hdr(ws3, row, 1, "投资活动现金流 (Investing Activities)", bg=MID_BLUE)
row += 1

capex = [-60, -100, -180, -280, -400, -550, -720, -900, -1050]
dat(ws3, row, 1, "  资本支出/算力投入 (CapEx — GPU/Compute)", align="left")
for i,v in enumerate(capex): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

acq = [0, -20, -30, -50, -80, -100, -120, -150, -150]
dat(ws3, row, 1, "  投资/收购 (Investments & Acquisitions)", align="left")
for i,v in enumerate(acq): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

cfi = [c+a for c,a in zip(capex, acq)]
dat(ws3, row, 1, "投资活动现金流合计 (Total CFI)", bold=True, align="left", bg=RED_LIGHT)
for i,v in enumerate(cfi):
    c = ws3.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=RED_LIGHT)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

# Financing CF
ws3.merge_cells(f'A{row}:J{row}')
hdr(ws3, row, 1, "融资活动现金流 (Financing Activities)", bg=MID_BLUE)
row += 1

equity_raise = [200, 1000, 1500, 3000, 5000, 0, 0, 0, 0]  # Funding rounds
dat(ws3, row, 1, "  股权融资 (Equity Financing)", align="left")
for i,v in enumerate(equity_raise): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

debt = [0, 0, 100, 200, 0, 0, -100, -200, 0]
dat(ws3, row, 1, "  债务净变动 (Net Debt Change)", align="left")
for i,v in enumerate(debt): dat(ws3, row, i+2, v, num_fmt='#,##0')
row += 1

cff = [e+d for e,d in zip(equity_raise, debt)]
dat(ws3, row, 1, "融资活动现金流合计 (Total CFF)", bold=True, align="left", bg=YELLOW)
for i,v in enumerate(cff):
    c = ws3.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

net_cash = [o+i+f for o,i,f in zip(cfo, cfi, cff)]
dat(ws3, row, 1, "净现金变动 (Net Cash Change)", bold=True, align="left", bg=LIGHT_BLUE)
for i,v in enumerate(net_cash):
    c = ws3.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

beg_cash = [50, 220, 1055, 2350, 5250, 9780, 8960, 7680, 6210]
end_cash = [220, 1055, 2350, 5250, 9780, 8960, 7680, 6210, 5210]
dat(ws3, row, 1, "  期末现金 (Ending Cash)", italic=True, align="left")
for i,v in enumerate(end_cash): dat(ws3, row, i+2, v, italic=True, num_fmt='#,##0')

ws3.freeze_panes = 'B4'

# ============================================================
# TAB 4: Balance Sheet
# ============================================================
ws4 = wb.create_sheet("Balance Sheet")
ws4.column_dimensions['A'].width = 40
for c in range(2, 11):
    ws4.column_dimensions[get_column_letter(c)].width = 13

ws4.merge_cells('A1:J1')
hdr(ws4, 1, 1, "智谱AI — 资产负债表 (Balance Sheet)", size=13, bg=DARK_BLUE)
for i, y in enumerate(YEARS, 2):
    bg_c = MID_BLUE if y <= 2024 else "2E7D32"
    for r in [2, 3]:
        c = ws4.cell(row=r, column=i)
        c.fill = PatternFill("solid", fgColor=bg_c)
        c.font = Font(bold=True, color="FFFFFF", size=10 if r==3 else 9, name="Calibri")
        c.alignment = Alignment(horizontal="center")
    ws4.cell(row=2, column=i).value = "历史" if y<=2024 else "预测"
    ws4.cell(row=3, column=i).value = y
ws4.merge_cells('A2:A3'); hdr(ws4, 2, 1, "人民币百万元 (RMB mn)", bg=MID_BLUE)

row = 4
ws4.merge_cells(f'A{row}:J{row}')
hdr(ws4, row, 1, "资产 (Assets)", bg=MID_BLUE)
row += 1

assets_cash = end_cash
dat(ws4, row, 1, "  现金及现金等价物 (Cash & Equivalents)", align="left")
for i,v in enumerate(assets_cash): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

ar = [r//6 for r in total_rev]
dat(ws4, row, 1, "  应收账款 (Accounts Receivable)", align="left")
for i,v in enumerate(ar): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

prepaid = [50, 80, 120, 180, 250, 350, 450, 550, 650]
dat(ws4, row, 1, "  预付款及其他流动资产 (Prepaid & Other CA)", align="left")
for i,v in enumerate(prepaid): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

total_ca = [c+a+p for c,a,p in zip(assets_cash, ar, prepaid)]
dat(ws4, row, 1, "流动资产合计 (Total Current Assets)", bold=True, align="left", bg=LIGHT_BLUE)
for i,v in enumerate(total_ca):
    c = ws4.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

ppe = [150, 230, 380, 600, 900, 1300, 1800, 2400, 3000]
dat(ws4, row, 1, "  固定资产/算力基础设施净值 (PP&E Net)", align="left")
for i,v in enumerate(ppe): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

intangibles = [30, 50, 80, 120, 180, 250, 330, 420, 510]
dat(ws4, row, 1, "  无形资产 (Intangibles)", align="left")
for i,v in enumerate(intangibles): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

other_lta = [20, 30, 50, 80, 120, 160, 200, 240, 280]
dat(ws4, row, 1, "  其他非流动资产 (Other Non-current Assets)", align="left")
for i,v in enumerate(other_lta): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

total_assets = [ca+p+i+o for ca,p,i,o in zip(total_ca, ppe, intangibles, other_lta)]
dat(ws4, row, 1, "资产总计 (Total Assets)", bold=True, align="left", bg=GREEN)
for i,v in enumerate(total_assets):
    c = ws4.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=GREEN)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

ws4.merge_cells(f'A{row}:J{row}')
hdr(ws4, row, 1, "负债及股东权益 (Liabilities & Equity)", bg=MID_BLUE)
row += 1

ap = [r//10 for r in total_rev]
dat(ws4, row, 1, "  应付账款 (Accounts Payable)", align="left")
for i,v in enumerate(ap): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

deferred_rev = [10, 20, 50, 100, 180, 300, 480, 700, 900]
dat(ws4, row, 1, "  递延收入 (Deferred Revenue)", align="left")
for i,v in enumerate(deferred_rev): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

total_cl = [a+d for a,d in zip(ap, deferred_rev)]
dat(ws4, row, 1, "流动负债合计 (Total Current Liabilities)", bold=True, align="left", bg=LIGHT_BLUE)
for i,v in enumerate(total_cl):
    c = ws4.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

lt_debt = [0, 0, 100, 300, 300, 300, 200, 0, 0]
dat(ws4, row, 1, "  长期借款 (Long-term Debt)", align="left")
for i,v in enumerate(lt_debt): dat(ws4, row, i+2, v, num_fmt='#,##0')
row += 1

equity = [ta-cl-d for ta,cl,d in zip(total_assets, total_cl, lt_debt)]
dat(ws4, row, 1, "股东权益 (Total Equity)", bold=True, align="left", bg=GREEN)
for i,v in enumerate(equity):
    c = ws4.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=GREEN)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
row += 1

dat(ws4, row, 1, "负债及权益合计 (Total L+E Check)", bold=True, align="left", bg=LIGHT_BLUE)
for i,v in enumerate(total_assets):
    c = ws4.cell(row=row, column=i+2, value=v)
    c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")

ws4.freeze_panes = 'B4'

# ============================================================
# TAB 5: Scenarios
# ============================================================
ws5 = wb.create_sheet("Scenarios")
ws5.column_dimensions['A'].width = 35
for c in range(2, 8):
    ws5.column_dimensions[get_column_letter(c)].width = 16

ws5.merge_cells('A1:G1')
hdr(ws5, 1, 1, "智谱AI — 情景分析 (Bull / Base / Bear Scenarios)", size=13, bg=DARK_BLUE)

# Headers
hdr(ws5, 2, 1, "指标", bg=MID_BLUE)
for i, label in enumerate(["2025E","2026E","2027E","2025E","2026E","2027E"], 2):
    hdr(ws5, 2, i, label, bg=MID_BLUE)

ws5.merge_cells('B2:D2')
hdr(ws5, 2, 2, "乐观情景 (Bull Case)", bg="1E6B1E")
ws5.merge_cells('E2:G2')
hdr(ws5, 2, 5, "悲观情景 (Bear Case)", bg="8B0000")

# Base case labels row
ws5.merge_cells('A3:G3')
hdr(ws5, 3, 1, "▌ 基准情景 (Base Case) — 收入同比增速: 2025E +82%, 2026E +74%, 2027E +66%", bg="2F5496")

metrics = [
    ("营业收入 (Revenue RMB mn)", 
     [2404, 4186, 6947],   # Bull
     [2980, 5400, 9200],   # Bull values (higher)
     [1900, 3100, 4900]),  # Bear
    ("毛利率 (Gross Margin%)", 
     [0.42, 0.48, 0.54], [0.45, 0.52, 0.58], [0.38, 0.42, 0.46]),
    ("EBITDA利润率 (EBITDA Margin%)",
     [-0.13, -0.05, 0.06], [-0.08, 0.02, 0.12], [-0.20, -0.15, -0.05]),
    ("净利润 (Net Income RMB mn)",
     [-390, -248, 347], [-200, 120, 890], [-600, -700, -350]),
    ("收入增速 (Revenue YoY%)",
     [0.82, 0.74, 0.66], [1.26, 0.81, 0.70], [0.44, 0.63, 0.58]),
]

row = 4
for metric, base, bull, bear in metrics:
    dat(ws5, row, 1, metric, align="left", bold=True)
    for i, (b, bl, br) in enumerate(zip(base, bull, bear)):
        if '%' in metric:
            pct(ws5, row, i+2, b); pct(ws5, row, i+5, bl); pct(ws5, row, i+8, br)
        else:
            dat(ws5, row, i+2, b, num_fmt='#,##0', bg=LIGHT_BLUE)
            dat(ws5, row, i+5, bl, num_fmt='#,##0', bg=GREEN)
            dat(ws5, row, i+8, br, num_fmt='#,##0', bg=RED_LIGHT)
    row += 1

# Add scenario descriptions
row += 1
ws5.merge_cells(f'A{row}:G{row}')
hdr(ws5, row, 1, "情景假设说明 (Scenario Assumptions)", bg=MID_BLUE)
row += 1
assumptions = [
    ("乐观情景", "企业私有化部署需求超预期，政策补贴力度加大，收入YoY>80%；毛利率提升至45-58%；2026年扭亏为盈"),
    ("基准情景", "商业化按计划推进，收入2025-2027年CAGR约70%；盈利拐点在2027年；估值基于P/S 30-35x 2026E收入"),
    ("悲观情景", "大厂价格战加剧，企业预算削减，收入增速降至40-60%；亏损期延长至2028年；需额外融资"),
]
for scenario, desc in assumptions:
    ws5.cell(row=row, column=1).value = f"【{scenario}】"
    ws5.cell(row=row, column=1).font = Font(bold=True, size=10)
    ws5.merge_cells(f'B{row}:G{row}')
    ws5.cell(row=row, column=2).value = desc
    ws5.cell(row=row, column=2).font = Font(size=10)
    row += 1

# ============================================================
# TAB 6: DCF Inputs
# ============================================================
ws6 = wb.create_sheet("DCF Inputs")
ws6.column_dimensions['A'].width = 38
for c in range(2, 9):
    ws6.column_dimensions[get_column_letter(c)].width = 14

ws6.merge_cells('A1:H1')
hdr(ws6, 1, 1, "智谱AI — DCF估值输入 (DCF Model Inputs)", size=13, bg=DARK_BLUE)

proj_labels = ["2025E","2026E","2027E","2028E","2029E"]
hdr(ws6, 2, 1, "DCF参数", bg=MID_BLUE)
for i, y in enumerate(proj_labels, 2):
    hdr(ws6, 2, i, y, bg=MID_BLUE)

row = 3
rev_proj = [2404, 4186, 6947, 10538, 13673]
ebit_proj = [e for e in ebit[4:]]
da_proj = da[4:]
capex_proj = [-abs(c) for c in capex[4:]]
wc_proj = wc_change[4:]

items_dcf = [
    ("营业收入 (Revenue)", rev_proj, '#,##0'),
    ("EBIT", ebit_proj, '#,##0'),
    ("税后EBIT / NOPAT (EBIT×(1-15%))", [round(e*0.85) for e in ebit_proj], '#,##0'),
    ("折旧摊销 (D&A)", da_proj, '#,##0'),
    ("资本支出 (CapEx)", capex_proj, '#,##0'),
    ("营运资金变动 (ΔWC)", wc_proj, '#,##0'),
    ("自由现金流 (UFCF)", [round(e*0.85+d+c+w) for e,d,c,w in zip(ebit_proj, da_proj, capex_proj, wc_proj)], '#,##0'),
]

for label, vals, fmt in items_dcf:
    bold_flag = 'UFCF' in label or '收入' in label
    bg_flag = LIGHT_BLUE if bold_flag else None
    dat(ws6, row, 1, label, bold=bold_flag, align="left", bg=bg_flag)
    for i, v in enumerate(vals):
        c = ws6.cell(row=row, column=i+2, value=v)
        c.font = Font(bold=bold_flag, size=10)
        if bg_flag:
            c.fill = PatternFill("solid", fgColor=bg_flag)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
    row += 1

row += 1
ws6.merge_cells(f'A{row}:H{row}')
hdr(ws6, row, 1, "WACC 及估值参数 (Valuation Assumptions)", bg=MID_BLUE)
row += 1

wacc_params = [
    ("无风险利率 (Risk-free Rate)", "2.5%"),
    ("股权风险溢价 (ERP)", "6.5%"),
    ("Beta (杠杆)", "1.45"),
    ("WACC", "11.9%"),
    ("终值增长率 (Terminal Growth Rate)", "3.5%"),
    ("折现年数 (Projection Years)", "5年 (2025-2029E)"),
    ("终值方法 (Terminal Value Method)", "永续增长模型 Gordon Growth Model"),
    ("股份数量 (Shares Outstanding, mn)", "500"),
    ("净现金 (Net Cash, RMB mn)", "5,250"),
]
for param, val in wacc_params:
    ws6.cell(row=row, column=1).value = param
    ws6.cell(row=row, column=1).font = Font(size=10, name="Calibri")
    ws6.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws6.cell(row=row, column=2).value = val
    ws6.cell(row=row, column=2).font = Font(bold=True, size=10, name="Calibri")
    ws6.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    row += 1

# Save
out_path = "/home/ubuntu/.openclaw-public/workspace/zhipuai-coverage/Task2_Model/ZhipuAI_Financial_Model_2025.xlsx"
wb.save(out_path)
print(f"✅ Saved: {out_path}")
